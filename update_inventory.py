#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to update Inventory BMT Excel file:
1. Insert STATUS column in position B
2. Add dropdown validation (ACTIVE, PARTIAL, USED)
3. Add conditional formatting (No fill, Yellow, Red)
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os
import shutil

# File paths
ORIGINAL_FILE = r"d:\BMT EXCEL\Inventory_bmt.xlsx"
BACKUP_FILE = r"d:\BMT EXCEL\Inventory_bmt_BACKUP.xlsx"
OUTPUT_FILE = r"d:\BMT EXCEL\Inventory_bmt_UPDATED.xlsx"

print("=" * 60)
print("BMT INVENTORY - STATUS COLUMN UPDATE SCRIPT")
print("=" * 60)

# Create backup
print("\n[1/6] Creating backup...")
shutil.copy2(ORIGINAL_FILE, BACKUP_FILE)
print(f"✓ Backup created: {BACKUP_FILE}")

# Load workbook
print("\n[2/6] Loading workbook...")
wb = load_workbook(ORIGINAL_FILE)
print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets")

# Define sheets to modify
sheet_names = ['ST.STEEL', 'STEEL', 'GALVANIZED', 'ZINCOR', 'ALUMINIUM',
               'BRASS', 'COPPER', 'ALUMINIUM PPTD', 'PLEXI', 'SKINPLATE', 'ACP']

print(f"\n[3/6] Processing {len(sheet_names)} sheets...")

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ Sheet '{sheet_name}' not found, skipping...")
        continue

    ws = wb[sheet_name]
    print(f"\n  📄 Processing: {sheet_name}")

    # Step 1: Insert new column at position B
    print("     → Inserting STATUS column at position B...")
    ws.insert_cols(2)  # Insert at column B (position 2)

    # Step 2: Update headers
    print("     → Updating headers...")
    ws['B1'] = 'STATUS'  # Row 1 header
    ws['B2'] = 'STATUS'  # Row 2 header

    # Apply header formatting to B1 and B2
    # Row 1 style (large, bold, blue background)
    ws['B1'].font = Font(name='Calibri', size=26, bold=True, color='FFFFFF')
    ws['B1'].fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Row 2 style (header, bold, darker blue)
    ws['B2'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['B2'].fill = PatternFill(start_color='2E75B5', end_color='2E75B5', fill_type='solid')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Set column width for STATUS column
    ws.column_dimensions['B'].width = 18

    # Step 3: Add Data Validation (Dropdown) for STATUS column
    print("     → Adding dropdown validation...")
    dv = DataValidation(
        type="list",
        formula1='"ACTIVE,PARTIAL,USED"',
        allow_blank=True
    )
    dv.error = 'Invalid value'
    dv.errorTitle = 'Invalid STATUS'
    dv.prompt = 'Select: ACTIVE, PARTIAL, or USED'
    dv.promptTitle = 'STATUS Selection'

    # Apply to rows 3-1000 in column B
    ws.add_data_validation(dv)
    dv.add('B3:B1000')

    # Step 4: Format data cells (B3:B1000)
    print("     → Formatting data cells...")
    for row in range(3, 1001):
        cell = ws[f'B{row}']
        cell.font = Font(name='Calibri', size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Step 5: Add Conditional Formatting
    print("     → Adding conditional formatting (color coding)...")

    # Yellow fill for PARTIAL
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    rule_partial = CellIsRule(
        operator='equal',
        formula=['"PARTIAL"'],
        fill=yellow_fill
    )
    ws.conditional_formatting.add('B3:B1000', rule_partial)

    # Red fill for USED
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    red_font = Font(color='FFFFFF', bold=True)  # White text on red background
    rule_used = CellIsRule(
        operator='equal',
        formula=['"USED"'],
        fill=red_fill,
        font=red_font
    )
    ws.conditional_formatting.add('B3:B1000', rule_used)

    # Note: ACTIVE or empty cells will have no fill (default white background)

    print(f"     ✓ Completed: {sheet_name}")

print("\n[4/6] Saving updated workbook...")
wb.save(OUTPUT_FILE)
print(f"✓ Saved: {OUTPUT_FILE}")

print("\n[5/6] Verifying output file...")
if os.path.exists(OUTPUT_FILE):
    file_size = os.path.getsize(OUTPUT_FILE) / 1024  # KB
    print(f"✓ File created successfully ({file_size:.2f} KB)")
else:
    print("✗ Error: Output file not created")

print("\n[6/6] Summary:")
print("=" * 60)
print(f"✓ Original file: {ORIGINAL_FILE}")
print(f"✓ Backup file:   {BACKUP_FILE}")
print(f"✓ Updated file:  {OUTPUT_FILE}")
print("\n✓ STATUS column added at position B")
print("✓ Dropdown validation: ACTIVE, PARTIAL, USED")
print("✓ Color coding:")
print("   - ACTIVE or empty = No fill (white)")
print("   - PARTIAL = Yellow background")
print("   - USED = Red background + white text")
print("=" * 60)
print("\n✅ ALL DONE! Upload the UPDATED file to Google Drive.")
print()
